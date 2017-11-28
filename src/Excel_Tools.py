# -*- coding: utf-8 -*-
'''
Created on 25.04.2017

@author: qi11028
'''

import re
from openpyxl import load_workbook
import xlrd
import xlwt
import datetime

import Witness
import Gui_OSDirectory

class ExcelUtils(object):
    clsStrFilePathFul = r'F:\SunsNoteBook\TM Objekte\Projekte\Kafas\Git_Prjs\fas2018\workspace_ECU-TEST\UserPyModules\sun_EclipseWorkspace\sunKafasCodesGen\src\ARXMLs\390\KOM_SP2018_17KW03_V9_USS_High_V8.xlsx'
    clsStrSheetName = u'Kommunikation'
    
    clsStrNumbersFilter1 = '\w[0-9]{8,20}\w'
    def __init__(self):
        pass

    def getListStrColumn(self, strExcelFileIn, intColumnIn, intMinRowIn, intMaxRowIn, strSheetNameIn= None):
        listStrOutput = []
        
        oWorkBook = load_workbook(filename=strExcelFileIn, read_only=True)
        if strSheetNameIn == None:
            oWorkSheet = oWorkBook.get_active_sheet()
        else:
            oWorkSheet = oWorkBook.get_sheet_by_name(strSheetNameIn)
            
        tuplesOftuplesRes = oWorkSheet.get_squared_range(min_col = intColumnIn, min_row = intMinRowIn, max_col = intColumnIn, max_row = intMaxRowIn)
        
        for tupleI in tuplesOftuplesRes:
            oCell = tupleI[0]
            listStrOutput.append(str(oCell.value))
        print (str(listStrOutput))
        return listStrOutput

    def filterListStrColumn_WitCol(self, listStr_ColumnIn, str_Re_SearchFilter_In):
        listStrOutput = []
        for tmpStrItem in listStr_ColumnIn:
            if re.search(str_Re_SearchFilter_In, tmpStrItem):
                listStrOutput.append(tmpStrItem)
        print (str(listStrOutput))
        return listStrOutput
              
    @staticmethod
    def get_listOpenpyxlRows_byExcelFile( strExcelFile_In, int_SheetIndex = None):
        '''
        this function is used to get a list of OpenpyxlRow objects with given excel file and a given sheet index
        relations: relations_1: this class, input: strExcelFile_In, Openpyxl
        relations: relations_2: this class, output: list of OpenpyxlRow objects

        @param strExcelFile_In: the excel file
        @type strExcelFile_In: string
        @param int_SheetIndex: sheet index of the excel file
        @type int_SheetIndex: integer
        @return: list of OpenpyxlRow objects, otherwise []
        @rtype: list of OpenpyxlRow objects
        '''
        strLocation = Witness.WitnessSys.clsStrWitnessLocation + "get_listOpenpyxlRows_byExcelFile: "
        listOutput = []
        oWorkbook1 = load_workbook(filename=strExcelFile_In, read_only=True)#
        oSheet = oWorkbook1.worksheets[int_SheetIndex]
            
        for row in oSheet.iter_rows():
            listOutput.append(row)
        print (strLocation + Witness.WitnessSys.clsStrWitnessValues + " listOutput size: " + str(len(listOutput)))     
        return listOutput
    

    
    @staticmethod
    def filter_Column_ListOpenpyxlRows( listObj_OpenpyxlRows_In, intColumn_In, str_RegexFilter_In):
        strLocation = Witness.WitnessSys.clsStrWitnessLocation + "filter_Column_ListOpenpyxlRows: "
        listOutput = []

        for row in listObj_OpenpyxlRows_In:
            if re.match(pattern = str_RegexFilter_In, string = str(row[intColumn_In].value)):
                listOutput.append(row)
        print (strLocation + Witness.WitnessSys.clsStrWitnessValues + " listOutput size: " + str(len(listOutput)))    
        return listOutput
    
    
    
    @staticmethod
    def filter_Column_ListOpenpyxlRows_byListTupleFilters( listObj_OpenpyxlRows_In, listTuple_RegexFilters_In):
        strLocation = Witness.WitnessSys.clsStrWitnessLocation + "filter_Column_ListOpenpyxlRows_byListTupleFilters: "
        listOutput = listObj_OpenpyxlRows_In

        for tuple_RegexColumFilter in listTuple_RegexFilters_In:
            listOutput = ExcelUtils.filter_Column_ListOpenpyxlRows(listObj_OpenpyxlRows_In = listOutput, 
                                                                   intColumn_In = tuple_RegexColumFilter[0], 
                                                                   str_RegexFilter_In = tuple_RegexColumFilter[1]
                                                                   )
        return listOutput
    
    @staticmethod
    def get_ListStr_Column_byListOpenpyxlRows( listObj_OpenpyxlRows_In, intColumn_In):
        strLocation = Witness.WitnessSys.clsStrWitnessLocation + "get_ListStr_Column_byListOpenpyxlRows: "
        listOutput = []
        for row in listObj_OpenpyxlRows_In:
            listOutput.append(str(row[intColumn_In].value))
        return listOutput
    
    @staticmethod
    def match_ListIndex_withListStrKeys_In_List( ListObj_In, ListStrKeys_RegexFilter_In):
        strLocation = Witness.WitnessSys.clsStrWitnessLocation + "match_ListIndex_withListStrKeys_In_List: "

        for strKeyFilter in ListStrKeys_RegexFilter_In:
            for listItem in ListObj_In:
                if re.match(pattern = strKeyFilter, string = str(listItem)):
                    return ListObj_In.index(listItem)
        print (strLocation + Witness.WitnessSys.clsStrWitnessValues + " error event: not found index" )    
        return None
    
class Xlrd_Utils(object):
    
    @staticmethod
    def get_listXlrdRows_byExcelFile( strExcelFile_In, int_SheetIndex = None):
        '''
        this function is used to get a list of XlrdRow objects with given excel file and a given sheet index
        relations: relations_1: this class, input: strExcelFile_In, Xlrd
        relations: relations_2: this class, output: list of XlrdRow objects

        @param strExcelFile_In: the excel file
        @type strExcelFile_In: string
        @param int_SheetIndex: sheet index of the excel file
        @type int_SheetIndex: integer
        @return: list of XlrdRow objects, otherwise []
        @rtype: list of XlrdRow objects
        '''
        strLocation = Witness.WitnessSys.clsStrWitnessLocation + "get_listXlrdRows_byExcelFile: "
        list_list_Output = []
        oWorkbook1 = xlrd.open_workbook(filename=strExcelFile_In)#
        oSheet = oWorkbook1.sheet_by_index(int_SheetIndex)
            
        for I_row in range(oSheet.nrows):
            list_list_Output.append(oSheet.row_values(I_row))
        print (strLocation + Witness.WitnessSys.clsStrWitnessValues + " listOutput size: " + str(len(list_list_Output)))     
        return list_list_Output
    
    @staticmethod
    def filter_Column_ListRows( list2_Rows_In, intColumn_In, str_RegexFilter_In):
        strLocation = Witness.WitnessSys.clsStrWitnessLocation + "filter_Column_ListOpenpyxlRows: "
        listOutput = []

        for list_row in list2_Rows_In:
            if re.match(pattern = str_RegexFilter_In, string = str(list_row[intColumn_In])):
                listOutput.append(list_row)
        print (strLocation + Witness.WitnessSys.clsStrWitnessValues + " listOutput size: " + str(len(listOutput)))    
        return listOutput
    
    @staticmethod
    def get_ListStr_Column_byList2Rows( list2_OpenpyxlRows_In, intColumn_In):
        strLocation = Witness.WitnessSys.clsStrWitnessLocation + "get_ListStr_Column_byListOpenpyxlRows: "
        listOutput = []
        for row in list2_OpenpyxlRows_In:
            listOutput.append(str(row[intColumn_In]))
        return listOutput
    
class BB_ExcelFileCalib_Data(object):

    def __init__(self, str_File_In, list_list_Rows_In):
        StrWitnessCurrent = Witness.WitnessSys.clsStrWitnessLocation + self.__class__.__name__+ ': __init__'
        self.str_File=str_File_In
        self.list2_Rows = list_list_Rows_In
        self.int_Ruag_Col = None
        self.int_Material_Col = None
        self.int_RCTotal_Col = None
        self.int_NRCTotal_Col = None
        self.int_LT_Col = None
        self.dict_Calib_Set = None
        self.list2_Rows_with_RowLimit = None
        
    def f_setCalibSet(self, list_RuId_In, list_Material_In, list_RC_In, list_NRC_In, list_LT_In, int_Row_Limit_In):
        self.dict_Calib_Set = {'list_RuIdRuid':list_RuId_In, 'list_Material':list_Material_In, 'list_RC':list_RC_In, 'list_NRC':list_NRC_In, 'list_LT':list_LT_In, "int_Row_Limit":int_Row_Limit_In}
        
    
    def f_calib_Row_Limit(self):
        StrWitnessCurrent = Witness.WitnessSys.clsStrWitnessLocation + self.__class__.__name__+ ': f_calib_Row_Limit: '
        if self.dict_Calib_Set == None:
            print(StrWitnessCurrent + "event: error: first run f_calib_Row_Limit")
            return 1
        if self.dict_Calib_Set['int_Row_Limit'] < len(self.list2_Rows):
            self.list2_Rows_with_RowLimit = self.list2_Rows[:self.dict_Calib_Set['int_Row_Limit']]
        else:
            self.list2_Rows_with_RowLimit = self.list2_Rows
        
    def f_calib_Ruid(self):
        StrWitnessCurrent = Witness.WitnessSys.clsStrWitnessLocation + self.__class__.__name__+ ': f_calib_Ruid: '
        if self.dict_Calib_Set == None:
            print(StrWitnessCurrent + "event: error: first run f_setCalibSet")
            raise UserWarning('CalibSet is not set')
        for listItem in self.list2_Rows_with_RowLimit:
            if set(listItem).intersection(self.dict_Calib_Set['list_RuIdRuid']):
                set_ColNames = set(listItem).intersection(self.dict_Calib_Set['list_RuIdRuid'])
                self.int_Ruag_Col = listItem.index(set_ColNames.pop())
                return 1
        raise UserWarning('f_calib_Ruid failed')
    
    def f_calib_Material(self):
        StrWitnessCurrent = Witness.WitnessSys.clsStrWitnessLocation + self.__class__.__name__+ ': f_calib_Material: '
        if self.dict_Calib_Set == None:
            print(StrWitnessCurrent + "event: error: first run f_setCalibSet")
            raise UserWarning('CalibSet is not set')
        for listItem in self.list2_Rows_with_RowLimit:
            if set(listItem).intersection(self.dict_Calib_Set['list_Material']):
                set_ColNames = set(listItem).intersection(self.dict_Calib_Set['list_Material'])
                self.int_Material_Col = listItem.index(set_ColNames.pop())
                return 1
        raise UserWarning('f_calib_Material failed')
    
    def f_calib_RC(self):
        StrWitnessCurrent = Witness.WitnessSys.clsStrWitnessLocation + self.__class__.__name__+ ': f_calib_RC: '
        if self.dict_Calib_Set == None:
            print(StrWitnessCurrent + "event: error: first run f_setCalibSet")
            raise UserWarning('CalibSet is not set')
        for listItem in self.list2_Rows_with_RowLimit:
            int_Index = ExcelUtils.match_ListIndex_withListStrKeys_In_List(ListObj_In = listItem, ListStrKeys_RegexFilter_In =  self.dict_Calib_Set['list_RC'])
            if int_Index:
               self.int_RCTotal_Col =  int_Index
               return 1
        raise UserWarning('f_calib_RC failed')
    
    def f_calib_NRC(self):
        StrWitnessCurrent = Witness.WitnessSys.clsStrWitnessLocation + self.__class__.__name__+ ': f_calib_NRC: '
        if self.dict_Calib_Set == None:
            print(StrWitnessCurrent + "event: error: first run f_setCalibSet")
            raise UserWarning('CalibSet is not set')
        for listItem in self.list2_Rows_with_RowLimit:
            if set(listItem).intersection(self.dict_Calib_Set['list_NRC']):
                set_ColNames = set(listItem).intersection(self.dict_Calib_Set['list_NRC'])
                self.int_NRCTotal_Col = listItem.index(set_ColNames.pop())
                return 1
        raise UserWarning('f_calib_NRC failed')
        
    def f_calib_LT(self):
        StrWitnessCurrent = Witness.WitnessSys.clsStrWitnessLocation + self.__class__.__name__+ ': f_calib_LT: '
        if self.dict_Calib_Set == None:
            print(StrWitnessCurrent + "event: error: first run f_setCalibSet")
            raise UserWarning('CalibSet is not set')
        for listItem in self.list2_Rows_with_RowLimit:
            if set(listItem).intersection(self.dict_Calib_Set['list_LT']):
                list_ColNames = set(listItem).intersection(self.dict_Calib_Set['list_LT'])
                self.int_LT_Col = listItem.index(list_ColNames[0])
                return 1
        raise UserWarning('f_calib_LT failed')
        
    def f_calib_File(self):
        self.f_calib_Ruid()
        self.f_calib_Material()
        self.f_calib_RC()
        self.f_calib_NRC()
        #self.f_calib_LT()
        
          
class BB_Excel_Data(object):
    clsStrFilePathFul = r'F:\tmp\xcels\E_oris\e1.xlsx'
    
    clsStr_Input_Dir = r'F:\tmp\xcels\E_oris\\'
    clsStr_Output_Dir = r'F:\tmp\xcels\Out\\'
    
    clsListStr_Ruag = ['RUAG Number']
    clsListStr_Material = ['Production material costs ']
    clsListStr_RC = ['Total RC . Part No.{0,10}']
    clsListStr_NRC = ['Total NRCs per Part No.']
    clsListStr_LT = []
    
    def __init__(self):
        StrWitnessCurrent = Witness.WitnessSys.clsStrWitnessLocation + self.__class__.__name__+ ': __init__'
        self.int_Material_Col = 11 
        self.int_Material_Col_Start = 11
        self.int_Material_Col_End = 12
        self.int_RCTotal_Col = 31
        self.int_NRCTotal_Col_Start = 33
        self.int_NRCTotal_Col_End = 35
        self.int_NRCTotal_Col = 37
        self.int_LT_Col_Start = 42
        self.int_LT_Col_End = 44
        
        self.list_tuple_PFile_Name_Rows = []
        
    def get_Obj_CompareData_byXlrdRow(self, obj_OpenpyxlRow_In ):
        strWitnessCurrent = Witness.WitnessSys.clsStrWitnessLocation + self.__class__.__name__+ "BB_Excel_Data:get_Obj_CompareData_byOpenpyxlRow: "
        listOutput_CompareItem = []
        listOutput_RuIdBlock = []
        listOutput_RuIdBlock = obj_OpenpyxlRow_In[1:6]
        str_Material=""
        float_NRC=0.0
        str_Material = obj_OpenpyxlRow_In[self.int_Material_Col_Start]
        str_RC = obj_OpenpyxlRow_In[self.int_RCTotal_Col]
        float_NRC = obj_OpenpyxlRow_In[self.int_NRCTotal_Col]
        str_LT = " "
        listOutput_CompareItem = [str_Material, str_RC, float_NRC, str_LT]
        print(strWitnessCurrent + "output: " + str(listOutput_CompareItem))
        return listOutput_RuIdBlock, listOutput_CompareItem
    
    def get_Obj_CompareData_byXlrdRow_1_1(self,  obj_OpenpyxlRow_In, Obj_BB_ExcelFileCalib_Data_In):
        strWitnessCurrent = Witness.WitnessSys.clsStrWitnessLocation + self.__class__.__name__+ "BB_Excel_Data:get_Obj_CompareData_byOpenpyxlRow: "
        listOutput_CompareItem = []
        listOutput_RuIdBlock = []
        listOutput_RuIdBlock = obj_OpenpyxlRow_In[1:6]
        str_Material=""
        float_NRC=0.0
        str_Material = obj_OpenpyxlRow_In[Obj_BB_ExcelFileCalib_Data_In.int_Material_Col]
        str_RC = obj_OpenpyxlRow_In[Obj_BB_ExcelFileCalib_Data_In.int_RCTotal_Col]
        float_NRC = obj_OpenpyxlRow_In[Obj_BB_ExcelFileCalib_Data_In.int_NRCTotal_Col]
        str_LT = " "
        listOutput_CompareItem = [str_Material, str_RC, float_NRC, str_LT]
        print(strWitnessCurrent + "output: " + str(listOutput_CompareItem))
        return listOutput_RuIdBlock, listOutput_CompareItem
    
    def get_list_list2_XlrdRows(self, str_InputDir_In = None):
        StrWitnessCurrent = Witness.WitnessSys.clsStrWitnessLocation + self.__class__.__name__+ ': get_list_list2_XlrdRows: ' 
        tuple_list_list2_XlrdRows = []
        if str_InputDir_In == None:
            str_InputDir_In = BB_Excel_Data.clsStr_Input_Dir
            
        listMatchedFiles = Gui_OSDirectory.OSDirectoryUtils.get_Dir_Files_with_Extention(str_InputDir_In.replace('\\', '/'), '.xlsx')
        for str_File in listMatchedFiles:
            print(StrWitnessCurrent + "Variable: str_File: " + str_File)
            list2Obj_001 = Xlrd_Utils.get_listXlrdRows_byExcelFile(strExcelFile_In = str_File, int_SheetIndex = 0)
            tuple_list_list2_XlrdRows.append((str_File, list2Obj_001))
        return tuple_list_list2_XlrdRows
    
    def build_BB_ExcelCalib_Data(self, list_tuple_list2_XlrdRows_In):
        listObj_Outputs = []
        for tuple_list2_XlrdRows in list_tuple_list2_XlrdRows_In:
            oBB_ExcelFileCalib_Data = BB_ExcelFileCalib_Data(str_File_In = tuple_list2_XlrdRows[0], list_list_Rows_In = tuple_list2_XlrdRows[1])
            oBB_ExcelFileCalib_Data.f_setCalibSet(list_RuId_In = BB_Excel_Data.clsListStr_Ruag, 
                                                  list_Material_In = BB_Excel_Data.clsListStr_Material, 
                                                  list_RC_In = BB_Excel_Data.clsListStr_RC, 
                                                  list_NRC_In = BB_Excel_Data.clsListStr_NRC, 
                                                  list_LT_In = BB_Excel_Data.clsListStr_LT, 
                                                  int_Row_Limit_In = 20)
            oBB_ExcelFileCalib_Data.f_calib_Row_Limit()
            oBB_ExcelFileCalib_Data.f_calib_File()
            listObj_Outputs.append(oBB_ExcelFileCalib_Data)
        return listObj_Outputs
            
            
    def get_ListStr_Ruids(self, list_tuple_list2_XlrdRows_In):
        StrWitnessCurrent = Witness.WitnessSys.clsStrWitnessLocation + self.__class__.__name__+ ': get_ListStr_Ruids: ' 
        
        list_RuIds = []

        for tuple_listXlrdRows in list_tuple_list2_XlrdRows_In:
            listObj_001 = Xlrd_Utils.filter_Column_ListRows(list2_Rows_In = tuple_listXlrdRows[1], intColumn_In = 2, str_RegexFilter_In = ExcelUtils.clsStrNumbersFilter1)
            listObj_001 = Xlrd_Utils.get_ListStr_Column_byList2Rows(list2_OpenpyxlRows_In = listObj_001, intColumn_In = 2)
            list_RuIds = list_RuIds + listObj_001
        list_RuIds = list(set(list_RuIds))
        return list_RuIds
    
    def get_list_Obj_Compare_TransferData_ByRuids(self,listStr_Ruids_In, list_tuple_list2_XlrdRows_In):
        list_Obj_Compare_TransferData = []
        for str_RuId in listStr_Ruids_In:
            oCompare_TransferData=Compare_TransferData(str_RuId)
            for tuple_list2_XlrdRows in list_tuple_list2_XlrdRows_In:
                listObj_001 = Xlrd_Utils.filter_Column_ListRows(list2_Rows_In = tuple_list2_XlrdRows[1], intColumn_In = 2, str_RegexFilter_In = str_RuId)
                if len(listObj_001)!=0:
                    print("Witness: Variable: str_RuId: " +str_RuId)
                    list_Str_Res_IdBlock, list_Str_Res_Compare = oBB_Excel_Data.get_Obj_CompareData_byXlrdRow(obj_OpenpyxlRow_In = listObj_001[0])
                    list_Str_Res_Compare.append(tuple_list2_XlrdRows[0])
                    if oCompare_TransferData.list_RuIdBlock == [] and list_Str_Res_IdBlock != []:
                        oCompare_TransferData.list_RuIdBlock = list_Str_Res_IdBlock
                    oCompare_TransferData.list2_CompareItem.append(list_Str_Res_Compare)
                else:
                    #len(listObj_001)==0
                    oCompare_TransferData.list2_CompareItem.append(['na.', 'na.', 'na.', 'na.'])
            list_Obj_Compare_TransferData.append(oCompare_TransferData)
        return list_Obj_Compare_TransferData
    
    def get_list_Obj_Compare_TransferData_ByRuids_1_1(self,listStr_Ruids_In, listObj_BB_ExcelFileCalib_Data_In):
        list_Obj_Compare_TransferData = []
        for str_RuId in listStr_Ruids_In:
            oCompare_TransferData=Compare_TransferData(str_RuId)
            for Obj_BB_ExcelFileCalib_Data in listObj_BB_ExcelFileCalib_Data_In:
                listObj_001 = Xlrd_Utils.filter_Column_ListRows(list2_Rows_In = Obj_BB_ExcelFileCalib_Data.list2_Rows, intColumn_In = Obj_BB_ExcelFileCalib_Data.int_Ruag_Col, str_RegexFilter_In = str_RuId)
                if len(listObj_001)!=0:
                    print("Witness: Variable: str_RuId: " +str_RuId)
                    list_Str_Res_IdBlock, list_Str_Res_Compare = oBB_Excel_Data.get_Obj_CompareData_byXlrdRow_1_1(obj_OpenpyxlRow_In = listObj_001[0], Obj_BB_ExcelFileCalib_Data_In = Obj_BB_ExcelFileCalib_Data)
                    list_Str_Res_Compare.append(Obj_BB_ExcelFileCalib_Data.str_File)
                    if oCompare_TransferData.list_RuIdBlock == [] and list_Str_Res_IdBlock != []:
                        oCompare_TransferData.list_RuIdBlock = list_Str_Res_IdBlock
                    oCompare_TransferData.list2_CompareItem.append(list_Str_Res_Compare)
                else:
                    #len(listObj_001)==0
                    oCompare_TransferData.list2_CompareItem.append(['na.', 'na.', 'na.', 'na.'])
            list_Obj_Compare_TransferData.append(oCompare_TransferData)
        return list_Obj_Compare_TransferData
    
    def save_get_list_Obj_Compare_TransferData_To_OutputDir(self, list_Obj_Compare_TransferData_In):
        oWT_Workbook = xlwt.Workbook()
        o_sheet1 = oWT_Workbook.add_sheet("Sheet1", cell_overwrite_ok=True)
        self.func_usecase_writeCompare_TransferData_ToExcel(list_Obj_Compare_TransferData_In = list_Obj_Compare_TransferData_In, o_xlwt_sheet_InOut = o_sheet1, int_Row_Start0_In = 6)
        str_File_Name_Pref = datetime.datetime.now().strftime('%Y_%m_%d_%H_%M_%S')
        oWT_Workbook.save(BB_Excel_Data.clsStr_Output_Dir.replace('\\', '/') + str_File_Name_Pref + "_out.xls")
    
    def func_usecase_writeCompare_TransferData_ToExcel(self, list_Obj_Compare_TransferData_In, o_xlwt_sheet_InOut, int_Row_Start0_In):
        for row_index, Obj_Compare_TransferData in enumerate(list_Obj_Compare_TransferData_In):
            o_xlwt_sheet_InOut.write(row_index + int_Row_Start0_In, 0, "")
            Xlwt_Utils.writListToRow(oSheet_InOut = o_xlwt_sheet_InOut, row_index_In = row_index + int_Row_Start0_In, col_index_Start = 1, list_toWrite_In = Obj_Compare_TransferData.list_RuIdBlock)
            int_LastCol_Index = 9
            for list_CompareItems in Obj_Compare_TransferData.list2_CompareItem:
                int_LastCol_Index = Xlwt_Utils.writListToRow(oSheet_InOut = o_xlwt_sheet_InOut, row_index_In = row_index + int_Row_Start0_In, col_index_Start = int_LastCol_Index, list_toWrite_In = list_CompareItems)
                int_LastCol_Index = int_LastCol_Index + 1 
    
class Compare_TransferData(object):
    
    def __init__(self, str_RuId_In):
        self.str_RuId = str_RuId_In
        self.list_RuIdBlock=[]
        self.list2_CompareItem=[]
        


class Xlwt_Utils(object):
    clsStr_ClassLocation = "Xlwt_Utils: "
    @staticmethod
    def writListToRow( oSheet_InOut, row_index_In, col_index_Start, list_toWrite_In):
        strLocation = Witness.WitnessSys.clsStrWitnessLocation + Xlwt_Utils.clsStr_ClassLocation + "writListToRow: "
        for col_index, IdItem in enumerate(list_toWrite_In):
            oSheet_InOut.write(row_index_In, col_index+col_index_Start, IdItem)
            col_index_output = col_index+col_index_Start
        return col_index_output
    


    
        

if __name__ == '__main__':

    
    oBB_Excel_Data = BB_Excel_Data()
    list_tuple_list2_XlrdRows = oBB_Excel_Data.get_list_list2_XlrdRows()
    listObj_BB_ExcelFileCalib_Data = oBB_Excel_Data.build_BB_ExcelCalib_Data(list_tuple_list2_XlrdRows_In = list_tuple_list2_XlrdRows)
    list_RuIds = oBB_Excel_Data.get_ListStr_Ruids(list_tuple_list2_XlrdRows_In = list_tuple_list2_XlrdRows)
    list_Obj_Compare_TransferData = oBB_Excel_Data.get_list_Obj_Compare_TransferData_ByRuids_1_1(listStr_Ruids_In = list_RuIds, listObj_BB_ExcelFileCalib_Data_In = listObj_BB_ExcelFileCalib_Data)
    oBB_Excel_Data.save_get_list_Obj_Compare_TransferData_To_OutputDir(list_Obj_Compare_TransferData_In = list_Obj_Compare_TransferData)  


